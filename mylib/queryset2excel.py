import copy
import csv
from os import path
import os
import re

import openpyxl
from lxml.doctestcompare import strip

"""
import openpyxl
openpyxl.xml.lxml_available()
"""

# y_name(sheets[0])
#     print(sheet.title)
from django.core.files.storage import default_storage
from django.db.models import F
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from mylib.my_common import get_digitalocean_spaces_download_url
from stats.models import Export
from stats.serializers import BaseDynamicStatsSerializer

from django.conf import settings


def serializerErrorToRow(headers, ser_data):
    # ser_data=serializer.data
    return [ser_data.get(header) for header in headers]


def serializerToRow(headers, ser_data):
    # ser_data=serializer.data
    return [ser_data.get(header.get("value")) for header in headers]


# from openpyxl import load_workbook
# wb = load_workbook(filename='large_file.xlsx', read_only=True)
# ws = wb['big_data']
#
# for row in ws.rows:
#     for cell in row:
#         print(cell.value)
#
# # Close the workbook after reading
# wb.close()


def get_row_value_name(column_name):
    if column_name is None:
        return ""
    return strip(column_name).lower().replace(" ", "_")


def get_header_column_name(column_name):
    return column_name.replace("_", " ").title()


def get_row_value(row, index):
    # print(index,len(row))
    if index > len(row) - 1:
        return None
    value = row[index].value

    if value != None:
        try:
            return value.strip()
        except Exception as e:
            try:
                return str(int(value))
            except Exception as e:
                return value
    return value


def importExcelCsv(filename, headers_only=False, include_rows_count=False, import_id=None):
    print("Loading book. lxml available ={}".format(openpyxl.xml.lxml_available()))
    if not openpyxl.xml.lxml_available():
        print("Lxml not installed")
        return

    wb = load_workbook(filename=filename, read_only=True)
    # print("Done Loading book")
    sheetnames = wb.sheetnames
    sheets_headers = []
    for sheet in sheetnames:
        ws = wb[sheet]
        headers = []
        row_count = 0
        if headers_only:
            headers = [cell.value for cell in next(ws.rows)]
            rows_count = -1
            if include_rows_count:
                for _ in ws.rows:
                    rows_count += 1
            sheets_headers.append(
                {
                    "name": sheet,
                    "rows_count": rows_count,
                    "headers": [get_row_value_name(header) for header in headers],
                }
            )
        else:
            for row in ws.rows:
                if len(headers) == 0:
                    headers = [cell.value for cell in row]
                    yield {
                        "header_row": True,
                        "sheet": sheet,
                        "sheetnames": sheetnames,
                        "headers": [get_row_value_name(header) for header in headers],
                    }
                else:
                    parsed_row = {get_row_value_name(header): get_row_value(row, index) for index, header in enumerate(headers)}
                    yield {"header_row": False, "sheet": sheet, "row": parsed_row}
    wb.close()
    if headers_only:
        yield sheets_headers


def ensure_dir_or_create(dir):
    if not os.path.exists(dir):
        os.makedirs(dir)


def exportExcelSheetOptimized(export_id, iterator, headers=[], filename=None, name=None):
    if len(headers) == 0 or filename is None:
        print("No filename or headers provided")
        return

    if not openpyxl.xml.lxml_available():
        Export.objects.filter(id=export_id).update(status="F", errors="lxml not installed.")
        return

    wb = Workbook(write_only=True)
    ws = wb.create_sheet()
    counter = 0
    updateAfter = 5000

    ##Write headers
    ws.append([header["name"] for header in headers])

    for data in iterator:
        ws.append(serializerToRow(headers, data))
        if counter >= updateAfter:
            Export.objects.filter(id=export_id).update(exported_rows_count=F("exported_rows_count") + counter)
            counter = 0
        counter += 1

    Export.objects.filter(id=export_id).update(exported_rows_count=F("exported_rows_count") + counter, status="P")
    # print(Export.objects.get(id=export_id).exported_rows_count)

    exports_dir_name = "Exports"

    # ensure_dir_or_create(path.join(settings.MEDIA_ROOT, exports_dir_name))
    # file_path = path.join(settings.MEDIA_ROOT, exports_dir_name, "{}-{}.xlsx".format(filename, export_id))
    xp = Export.objects.get(id=export_id)

    file_path = f"Temp{xp.id}.xlsx"
    wb.save(file_path)

    export_path = path.join(f"Exports", "{}-{}.xlsx".format(filename, export_id))
    res = default_storage.save(export_path, open(file_path, "rb"))
    print(res)
    try:
        os.remove(file_path)
    except Exception as e:
        print(e)

    xp.finish(res)


def exportcsv(
    headers=[],
    title="Sheet",
    filename=None,
    queryset=[],
    export_csv=False,
    request=None,
):
    """
    Example
    filename="test"
    queryset=[{"school_title":"Warugara","count":4}]
    headers=[{"name":"School Title","value":"school_title"},{"name":"Students Count","value":"count"}]
    path=exportcsv(filename=filename,queryset=queryset,headers=headers,title="Schools",export_csv=True,request=None)
    :param headers:
    :param title:
    :param filename:
    :param querset:
    :return:
    """
    path = ""
    #####Validate data, assert headers count match an ojbects attributes

    ###Get the totals
    queryset_length = len(queryset)
    headers_length = len(headers)

    ##New workbook
    wb = openpyxl.Workbook()

    ##Create a sheet
    sheet = wb.active
    sheet.title = title

    ###Set the headers

    for k, col in enumerate(headers):
        cell = sheet.cell(row=1, column=k + 1)
        cell.value = col["name"]

        ####Set the size
        if k + 1 <= headers_length:
            sheet.column_dimensions[get_column_letter(k + 1)].width = 20

    ###Copy the headers to match the number of fields in data
    # try:
    fields_length = len([k for k in queryset[0]])
    # except:
    #     fields_length=len(k for k in queryset[0])
    ##Get number of attributes per  row

    myheaders = copy.deepcopy(headers[:fields_length])

    ####Writing the data
    for i, data in enumerate(queryset):
        ###Loop through all the headers
        for j, col in enumerate(myheaders):
            ##i+2 since (i starts at 0, and the row 1 is for headers)
            dt = data[col["value"]]
            # print(dt)
            dat = ",".join(list(dt)) if type(dt) in [list, set] else dt
            # print(dt,dat)
            sheet.cell(row=i + 2, column=j + 1).value = dat

    ##The output filename

    myfilename = "%s.%s" % (filename, "csv" if export_csv else ".xlsx")

    ####Temporatyfilename for openxl
    temp_filename = "temp_%s" % (myfilename)

    ##Temporarily save the file
    if export_csv:
        with open(temp_filename, "wb") as f:  # open('test.csv', 'w', newline="") for python 3
            c = csv.writer(f)
            for r in sheet.rows:
                c.writerow([cell.value for cell in r])
    else:
        wb.save(temp_filename)

    print(myfilename)
    with open(temp_filename) as f:
        default_storage.delete("{}".format(myfilename))
        path = default_storage.save("{}".format(myfilename), f)
    url = path
    if request:
        url = get_digitalocean_spaces_download_url(path)  # request.build_absolute_uri(location="/media/" + path)
    return url


if __name__ == "__main__":
    pass
